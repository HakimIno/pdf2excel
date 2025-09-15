#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF-Like Excel Writer
====================

Creates Excel files that look exactly like the original PDF
using the Intelligent PDF Reader system.
"""

import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_LIBS_AVAILABLE = True
except ImportError as e:
    EXCEL_LIBS_AVAILABLE = False
    print(f"Excel libraries not available: {e}")

try:
    from .table_filter import TableFilter
    TABLE_FILTER_AVAILABLE = True
except ImportError:
    TABLE_FILTER_AVAILABLE = False

try:
    from .intelligent_pdf_reader import IntelligentPDFReader
    INTELLIGENT_READER_AVAILABLE = True
except ImportError:
    INTELLIGENT_READER_AVAILABLE = False


class PDFLikeWriter:
    """
    Excel writer that creates output identical to PDF layout
    using intelligent PDF analysis without hard-coding
    """
    
    def __init__(self):
        """Initialize PDF-like writer"""
        self.logger = logging.getLogger(__name__)
        
        if not EXCEL_LIBS_AVAILABLE:
            self.logger.error("Excel libraries not available")
            
        # Initialize table filter
        if TABLE_FILTER_AVAILABLE:
            self.table_filter = TableFilter()
        else:
            self.table_filter = None
            
        # Initialize intelligent PDF reader (main system)
        if INTELLIGENT_READER_AVAILABLE:
            self.intelligent_reader = IntelligentPDFReader()
        else:
            self.intelligent_reader = None
            
    def write_to_excel(self, pdf_data: Dict[str, Any], output_path: str) -> bool:
        """
        Write PDF data to Excel with exact PDF layout
        
        Args:
            pdf_data (dict): Extracted PDF data
            output_path (str): Output Excel file path
            
        Returns:
            bool: True if successful
        """
        if not EXCEL_LIBS_AVAILABLE:
            return self._write_text_format(pdf_data, output_path)
            
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Get pages data
            pages_data = pdf_data.get('pages', [])
            if not pages_data:
                self.logger.warning("No page data found")
                return False
                
            # Create one sheet per page with intelligent layout
            for page_num, page_content in enumerate(pages_data, 1):
                self._create_intelligent_pdf_sheet(wb, page_num, page_content, pdf_data)
                
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"PDF-like Excel file created: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating PDF-like Excel: {str(e)}")
            return False
            
    def _create_intelligent_pdf_sheet(self, wb: Workbook, page_num: int, page_content: Dict[str, Any], 
                                    pdf_data: Dict[str, Any]) -> None:
        """Create intelligent PDF sheet using PyMuPDF analysis"""
        ws = wb.create_sheet(f"หน้า_{page_num}")
        
        # Use intelligent PDF reader (primary method)
        if self.intelligent_reader:
            pdf_path = pdf_data.get('pdf_path')
            if pdf_path:
                layout_info = self.intelligent_reader.analyze_pdf_layout(pdf_path)
                if layout_info:
                    self.intelligent_reader.create_excel_from_layout(ws, layout_info)
                    return
                    
        # Fallback to basic method if intelligent reader fails
        self._create_basic_sheet(ws, page_content, pdf_data, page_num)
        
    def _create_basic_sheet(self, ws, page_content: Dict[str, Any], pdf_data: Dict[str, Any], page_num: int) -> None:
        """Basic fallback method"""
        text_content = page_content.get('text_content', '')
        tables = page_content.get('tables', [])
        
        # Filter real tables
        if self.table_filter:
            tables = self.table_filter.filter_real_tables(tables)
        
        # Create simple layout
        current_row = 1
        
        # Add text content
        if text_content:
            lines = text_content.split('\n')
            for line in lines:
                line = line.strip()
                if line:
                    ws[f'A{current_row}'] = line
                    ws[f'A{current_row}'].font = Font(size=10)
                    current_row += 1
                
        # Add tables
        for table in tables:
            table_data = table.get('data', [])
            if table_data:
                current_row += 1  # Space before table
                
                # Add table using extracted colors
                page_colors = page_content.get('colors', {}).get('table_colors', {})
                
                # Use intelligent column layout for full width usage
                for row_idx, row_data in enumerate(table_data):
                    # Clean row data - remove empty cells
                    cleaned_row = [str(cell).strip() for cell in row_data if str(cell).strip()]
                    
                    if not cleaned_row:  # Skip completely empty rows
                        continue
                    
                    # Calculate optimal column distribution to span A-H
                    num_cols = len(cleaned_row)
                    # Use complex layout for better full-width utilization
                    col_positions = self._calculate_optimal_column_layout(num_cols)
                    
                    # Create cells with proper column distribution
                    for col_idx, cell_value in enumerate(cleaned_row):
                        if col_idx >= len(col_positions):
                            break
                            
                        start_col, end_col = col_positions[col_idx]
                        start_letter = get_column_letter(start_col)
                        end_letter = get_column_letter(end_col)
                        
                        if start_col == end_col:
                            # Single column
                            cell = ws[f'{start_letter}{current_row}']
                            cell.value = cell_value
                        else:
                            # Merged columns
                            ws.merge_cells(f'{start_letter}{current_row}:{end_letter}{current_row}')
                            cell = ws[f'{start_letter}{current_row}']
                            cell.value = cell_value
                        
                        # Apply formatting
                        if row_idx == 0:  # Header
                            header_text_color = page_colors.get('header_text')
                            header_bg_color = page_colors.get('header_bg')
                            
                            # Only apply colors if they exist in PDF
                            if header_text_color:
                                cell.font = Font(size=10, bold=True, color=header_text_color)
                            else:
                                cell.font = Font(size=10, bold=True)  # Default black text
                                
                            if header_bg_color:
                                cell.fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
                            # No fill if no background color in PDF
                        else:
                            data_text_color = page_colors.get('data_text')
                            
                            # Only apply text color if it exists in PDF
                            if data_text_color:
                                cell.font = Font(size=10, color=data_text_color)
                            else:
                                cell.font = Font(size=10)  # Default black text
                        
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Only add borders if border color exists in PDF
                        border_color = page_colors.get('border_color')
                        if border_color:
                            for col in range(start_col, end_col + 1):
                                border_cell = ws[f'{get_column_letter(col)}{current_row}']
                                border_cell.border = Border(
                                    left=Side(style='thin', color=border_color),
                                    right=Side(style='thin', color=border_color),
                                    top=Side(style='thin', color=border_color),
                                    bottom=Side(style='thin', color=border_color)
                                )
                current_row += 1
                
        # Set column widths to ensure full width usage - optimized for full width table display
        column_widths = [30, 35, 20, 25, 25, 25, 25, 25]  # Optimized widths for full width payroll statement
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _calculate_optimal_column_layout(self, num_cols: int) -> List[Tuple[int, int]]:
        """Calculate optimal column layout to span A-H efficiently - optimized for payroll statements"""
        if num_cols == 1:
            return [(1, 8)]  # A-H (full width)
        elif num_cols == 2:
            return [(1, 4), (5, 8)]  # A-D, E-H (half width each)
        elif num_cols == 3:
            return [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H (balanced)
        elif num_cols == 4:
            return [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
        elif num_cols == 5:
            # Optimized for payroll statements: Category, Deduction Type, Rate, Current, Year to Date
            return [(1, 1), (2, 3), (4, 4), (5, 6), (7, 8)]  # A, B-C, D, E-F, G-H
        elif num_cols == 6:
            return [(1, 1), (2, 2), (3, 3), (4, 4), (5, 6), (7, 8)]  # A, B, C, D, E-F, G-H
        elif num_cols == 7:
            return [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 8)]  # A, B, C, D, E, F, G-H
        elif num_cols == 8:
            return [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8)]  # A, B, C, D, E, F, G, H
        else:
            return [(i+1, i+1) for i in range(min(num_cols, 8))]
            
    def _write_text_format(self, pdf_data: Dict[str, Any], output_path: str) -> bool:
        """Fallback to text format if Excel not available"""
        try:
            with open(output_path.replace('.xlsx', '.txt'), 'w', encoding='utf-8') as f:
                f.write("PDF to Excel Conversion Results\n")
                f.write("=" * 50 + "\n\n")
                
                pages_data = pdf_data.get('pages', [])
                for page_num, page_content in enumerate(pages_data, 1):
                    f.write(f"Page {page_num}\n")
                    f.write("-" * 20 + "\n")
                    
                    text_content = page_content.get('text_content', '')
                    if text_content:
                        f.write(text_content)
                    f.write("\n\n")
                    
                    tables = page_content.get('tables', [])
                    for table_idx, table in enumerate(tables, 1):
                        f.write(f"Table {table_idx}:\n")
                        table_data = table.get('data', [])
                        for row in table_data:
                            f.write(" | ".join(str(cell) for cell in row))
                            f.write("\n")
                        f.write("\n")
                        
            self.logger.info(f"Text format file created: {output_path.replace('.xlsx', '.txt')}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating text format: {str(e)}")
            return False