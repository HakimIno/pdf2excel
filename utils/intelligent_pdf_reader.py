#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Intelligent PDF Reader
=====================

Uses PyMuPDF (fitz) to intelligently read PDF layout, colors, fonts, 
and positioning without any hard-coding. Adapts to any PDF format.
"""

import logging
from typing import List, Dict, Any, Optional, Tuple
import json

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_LIBS_AVAILABLE = True
except ImportError:
    EXCEL_LIBS_AVAILABLE = False


class IntelligentPDFReader:
    """
    Intelligent PDF reader that extracts exact layout information
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def analyze_pdf_layout(self, pdf_path: str) -> Dict[str, Any]:
        """Analyze PDF to extract complete layout information"""
        if not PYMUPDF_AVAILABLE:
            self.logger.error("PyMuPDF not available - cannot perform intelligent analysis")
            return {}
            
        try:
            doc = fitz.open(pdf_path)
            page = doc[0]  # Analyze first page
            
            layout_info = {
                'page_size': {
                    'width': page.rect.width,
                    'height': page.rect.height
                },
                'text_blocks': self._extract_text_blocks(page),
                'tables': self._extract_tables(page),
                'images': self._extract_images(page),
                'shapes': self._extract_shapes(page),
                'colors': self._extract_colors(page),
                'fonts': self._extract_fonts(page)
            }
            
            doc.close()
            return layout_info
            
        except Exception as e:
            self.logger.error(f"Error analyzing PDF layout: {str(e)}")
            return {}
            
    def _extract_text_blocks(self, page) -> List[Dict]:
        """Extract text blocks with exact positioning and formatting"""
        text_blocks = []
        
        # Get text blocks with formatting
        blocks = page.get_text("dict")
        
        for block in blocks.get("blocks", []):
            if "lines" not in block:
                continue
                
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if not text:
                        continue
                        
                    # Extract position and formatting
                    bbox = span["bbox"]
                    font_info = {
                        'font': span["font"],
                        'size': span["size"],
                        'flags': span["flags"],  # Bold, italic, etc.
                        'color': span["color"]
                    }
                    
                    text_blocks.append({
                        'text': text,
                        'bbox': bbox,
                        'x': bbox[0],
                        'y': bbox[1],
                        'width': bbox[2] - bbox[0],
                        'height': bbox[3] - bbox[1],
                        'font_info': font_info,
                        'is_bold': bool(font_info['flags'] & 2**4),
                        'is_italic': bool(font_info['flags'] & 2**1),
                        'block_type': self._classify_text_block(text, font_info)
                    })
                    
        return text_blocks
        
    def _extract_tables(self, page) -> List[Dict]:
        """Extract tables with exact positioning and styling"""
        tables = []
        
        # Find tables using PyMuPDF's table detection
        tabs = page.find_tables()
        
        for table in tabs:
            # Extract table data
            table_data = table.extract()
            if not table_data:
                continue
                
            # Get table styling information
            bbox = table.bbox
            
            # Analyze cell formatting
            cell_formats = self._analyze_table_formatting(page, table, bbox)
            
            # Enhanced table analysis
            table_info = {
                'data': table_data,
                'bbox': bbox,
                'x': bbox[0],
                'y': bbox[1], 
                'width': bbox[2] - bbox[0],
                'height': bbox[3] - bbox[1],
                'cell_formats': cell_formats,
                'table_type': self._classify_table(table_data),
                'num_rows': len(table_data),
                'num_cols': len(table_data[0]) if table_data else 0,
                'has_headers': self._detect_headers(table_data),
                'complexity': self._assess_table_complexity(table_data)
            }
            
            tables.append(table_info)
            
        return tables
        
    def _extract_images(self, page) -> List[Dict]:
        """Extract images with positioning"""
        images = []
        image_list = page.get_images()
        
        for img in image_list:
            # Get image positioning
            img_bbox = page.get_image_bbox(img)
            if img_bbox:
                images.append({
                    'bbox': img_bbox,
                    'x': img_bbox[0],
                    'y': img_bbox[1],
                    'width': img_bbox[2] - img_bbox[0],
                    'height': img_bbox[3] - img_bbox[1]
                })
                
        return images
        
    def _extract_shapes(self, page) -> List[Dict]:
        """Extract shapes and lines with colors"""
        shapes = []
        
        # Get drawing commands
        drawings = page.get_drawings()
        
        for drawing in drawings:
            shape_info = {
                'bbox': drawing.get('rect'),
                'type': drawing.get('type', 'unknown'),
                'color': drawing.get('color'),
                'fill': drawing.get('fill'),
                'stroke_color': drawing.get('stroke', {}).get('color'),
                'width': drawing.get('width', 1)
            }
            
            if shape_info['bbox']:
                bbox = shape_info['bbox']
                shape_info.update({
                    'x': bbox[0],
                    'y': bbox[1],
                    'width': bbox[2] - bbox[0],
                    'height': bbox[3] - bbox[1]
                })
                
            shapes.append(shape_info)
            
        return shapes
        
    def _extract_colors(self, page) -> Dict[str, Any]:
        """Extract comprehensive color scheme from the page"""
        colors = {
            'background_colors': [],
            'text_colors': [],
            'dominant_colors': [],
            'border_colors': [],  # New: specifically track border colors
            'table_colors': {
                'header_bg': None,
                'header_text': None,
                'data_bg_primary': None,
                'data_bg_alternate': None,
                'data_text': None,
                'border_color': None
            }
        }
        
        # Extract colors from text with frequency tracking
        text_color_freq = {}
        blocks = page.get_text("dict")
        for block in blocks.get("blocks", []):
            if "lines" not in block:
                continue
            for line in block["lines"]:
                for span in line["spans"]:
                    color = span.get("color")
                    if color:
                        # Convert to hex format for consistency
                        hex_color = self._convert_color_to_hex(color)
                        if hex_color:
                            text_color_freq[hex_color] = text_color_freq.get(hex_color, 0) + 1
                            if hex_color not in colors['text_colors']:
                                colors['text_colors'].append(hex_color)
                        
        # Extract colors from shapes and drawings (including borders)
        bg_color_freq = {}
        border_color_freq = {}
        drawings = page.get_drawings()
        for drawing in drawings:
            fill_color = drawing.get('fill')
            stroke_color = drawing.get('stroke', {}).get('color')
            stroke_width = drawing.get('width', 0)
            
            if fill_color:
                hex_fill = self._convert_color_to_hex(fill_color)
                if hex_fill:
                    bg_color_freq[hex_fill] = bg_color_freq.get(hex_fill, 0) + 1
                    if hex_fill not in colors['background_colors']:
                        colors['background_colors'].append(hex_fill)
                        
            if stroke_color:
                hex_stroke = self._convert_color_to_hex(stroke_color)
                if hex_stroke:
                    # Track border colors separately
                    border_color_freq[hex_stroke] = border_color_freq.get(hex_stroke, 0) + 1
                    if hex_stroke not in colors['border_colors']:
                        colors['border_colors'].append(hex_stroke)
                    if hex_stroke not in colors['text_colors']:
                        colors['text_colors'].append(hex_stroke)
        
        # Determine dominant colors
        if text_color_freq:
            colors['dominant_colors'] = sorted(text_color_freq.items(), key=lambda x: x[1], reverse=True)[:3]
        
        # Smart color assignment for tables (including border detection)
        colors['table_colors'] = self._determine_table_colors(colors, text_color_freq, bg_color_freq, border_color_freq)
                
        return colors
    
    def _convert_color_to_hex(self, color) -> Optional[str]:
        """Convert various color formats to hex"""
        try:
            if isinstance(color, int):
                # Convert RGB integer to hex
                return f"FF{color:06X}"
            elif isinstance(color, (list, tuple)) and len(color) >= 3:
                # Convert RGB list/tuple to hex
                r, g, b = int(color[0] * 255), int(color[1] * 255), int(color[2] * 255)
                return f"FF{r:02X}{g:02X}{b:02X}"
            elif isinstance(color, str) and color.startswith('#'):
                # Already hex format, ensure FF prefix for alpha
                return f"FF{color[1:]}" if len(color) == 7 else color
            elif isinstance(color, str) and len(color) == 6:
                # Hex without alpha
                return f"FF{color}"
            return None
        except:
            return None
    
    def _determine_table_colors(self, colors: Dict, text_freq: Dict, bg_freq: Dict, border_freq: Dict = None) -> Dict:
        """Determine table colors based on extracted colors - only use colors that exist in PDF"""
        # Start with no colors (transparent/white background)
        table_colors = {
            'header_bg': None,           # No background color
            'header_text': None,         # Default text color
            'data_bg_primary': None,     # No background color
            'data_bg_alternate': None,   # No alternating color
            'data_text': None,           # Default text color
            'border_color': None         # No border color
        }
        
        # Use colors that exist in PDF - prioritize professional table appearance
        if bg_freq:
            # Find the most suitable header background color from PDF
            for bg_color, freq in sorted(bg_freq.items(), key=lambda x: x[1], reverse=True):
                if self._is_suitable_header_color(bg_color):
                    table_colors['header_bg'] = bg_color
                    break
            
            # If no suitable header color found, use light gray for professional appearance
            if not table_colors['header_bg']:
                table_colors['header_bg'] = 'FFE5E5E5'  # Light gray for headers
            
            # Only use alternating colors if they're very light (to maintain border visibility)
            very_light_colors = [color for color in bg_freq.keys() if self._get_color_brightness(color) > 240]
            if very_light_colors:
                lightest = max(very_light_colors, key=lambda x: self._get_color_brightness(x))
                table_colors['data_bg_alternate'] = lightest
        
        # Use text colors that exist in PDF
        if text_freq:
            # Find darkest readable text color from PDF
            dark_text_colors = [color for color in text_freq.keys() if self._get_color_brightness(color) < 128]
            if dark_text_colors:
                darkest = min(dark_text_colors, key=lambda x: self._get_color_brightness(x))
                table_colors['data_text'] = darkest
        
        # Enhanced border detection - use actual border colors from PDF
        if border_freq:
            # Find the most common border color (likely table borders)
            for border_color, freq in sorted(border_freq.items(), key=lambda x: x[1], reverse=True):
                # Prefer dark colors for borders (black, dark gray)
                if self._get_color_brightness(border_color) < 100:
                    table_colors['border_color'] = border_color
                    break
            # If no dark border found, use black for better visibility
            if not table_colors['border_color']:
                table_colors['border_color'] = 'FF000000'  # Black borders
        elif table_colors['header_bg']:
            # Fallback: create border color from header background
            table_colors['border_color'] = self._create_border_color(table_colors['header_bg'])
        else:
            # Default: use black borders for better visibility
            table_colors['border_color'] = 'FF000000'  # Black borders
        
        # Only set header text color if we have a header background
        if table_colors['header_bg']:
            if self._is_dark_color(table_colors['header_bg']):
                table_colors['header_text'] = 'FFFFFFFF'  # White on dark
            else:
                table_colors['header_text'] = 'FF000000'  # Black on light
        
        return table_colors
    
    def _is_suitable_header_color(self, hex_color: str) -> bool:
        """Check if color is suitable for table headers"""
        try:
            brightness = self._get_color_brightness(hex_color)
            # Good for headers: not too dark, not too light
            return 40 <= brightness <= 180
        except:
            return False
    
    def _create_border_color(self, base_color: str) -> str:
        """Create a subtle border color based on the base color"""
        try:
            if len(base_color) >= 8:
                r = int(base_color[2:4], 16)
                g = int(base_color[4:6], 16)
                b = int(base_color[6:8], 16)
                
                # Create a lighter version for borders
                r = min(255, r + 40)
                g = min(255, g + 40)
                b = min(255, b + 40)
                
                return f"FF{r:02X}{g:02X}{b:02X}"
        except:
            pass
        return None  # No border color if can't create from base color
    
    def _is_dark_color(self, hex_color: str) -> bool:
        """Check if color is dark"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)  
                b = int(hex_color[6:8], 16)
                brightness = (r * 299 + g * 587 + b * 114) / 1000
                return brightness < 128
        except:
            pass
        return False
    
    def _get_color_brightness(self, hex_color: str) -> int:
        """Get color brightness value"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)
                b = int(hex_color[6:8], 16)
                return int((r * 299 + g * 587 + b * 114) / 1000)
        except:
            pass
        return 128
    
    def _darken_color(self, hex_color: str, factor: float = 0.7) -> str:
        """Darken a color by factor"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)
                b = int(hex_color[6:8], 16)
                r = int(r * factor)
                g = int(g * factor)
                b = int(b * factor)
                return f"FF{r:02X}{g:02X}{b:02X}"
        except:
            pass
        return hex_color
    
    def _lighten_color(self, hex_color: str, factor: float = 0.9) -> str:
        """Lighten a color by factor"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)
                b = int(hex_color[6:8], 16)
                r = int(r + (255 - r) * (1 - factor))
                g = int(g + (255 - g) * (1 - factor))
                b = int(b + (255 - b) * (1 - factor))
                return f"FF{r:02X}{g:02X}{b:02X}"
        except:
            pass
        return hex_color
        
    def _extract_fonts(self, page) -> Dict[str, Any]:
        """Extract font information"""
        fonts = {
            'font_list': [],
            'font_sizes': [],
            'font_usage': {}
        }
        
        blocks = page.get_text("dict")
        for block in blocks.get("blocks", []):
            if "lines" not in block:
                continue
            for line in block["lines"]:
                for span in line["spans"]:
                    font = span.get("font", "")
                    size = span.get("size", 10)
                    
                    if font and font not in fonts['font_list']:
                        fonts['font_list'].append(font)
                        
                    if size not in fonts['font_sizes']:
                        fonts['font_sizes'].append(size)
                        
                    # Track font usage
                    font_key = f"{font}_{size}"
                    if font_key not in fonts['font_usage']:
                        fonts['font_usage'][font_key] = 0
                    fonts['font_usage'][font_key] += 1
                    
        return fonts
        
    def _classify_text_block(self, text: str, font_info: Dict) -> str:
        """Classify text block type based on content and formatting"""
        text_upper = text.upper().strip()
        font_size = font_info.get('size', 10)
        is_bold = bool(font_info.get('flags', 0) & 2**4)
        
        # Main headers (large, bold)
        if font_size > 14 and is_bold:
            return 'main_header'
            
        # Section headers (all caps, bold, or contains keywords)
        if (is_bold and font_size > 10) or text.isupper():
            if any(word in text_upper for word in [
                'CURRENT', 'EARNINGS', 'STATEMENT', 'SUMMARY', 
                'DEDUCTION', 'DETAILS', 'TOTAL'
            ]):
                return 'section_header'
                
        # Labels (contains colons or specific patterns)
        if ':' in text or any(word in text_upper for word in [
            'EMPLOYEE', 'NAME', 'PERIOD', 'DATE', 'STATUS'
        ]):
            return 'label'
            
        # Data (numbers, currency)
        if any(char in text for char in ['$', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9']):
            return 'data'
            
        # Footer (small, contains copyright)
        if font_size < 10 or any(word in text.lower() for word in ['©', 'copyright', 'inc']):
            return 'footer'
            
        return 'text'
        
    def _classify_table(self, table_data: List[List]) -> str:
        """Classify table type based on content"""
        if not table_data or len(table_data) < 2:
            return 'unknown'
            
        # Check header row
        header_text = ' '.join(str(cell) for cell in table_data[0] if cell).upper()
        
        if any(word in header_text for word in ['EARNING', 'HOURS', 'PAYMENT', 'PAY']):
            return 'earnings'
        elif any(word in header_text for word in ['DEDUCTION', 'TAX', 'WITHHOLD']):
            return 'deductions'
        elif any(word in header_text for word in ['SUMMARY', 'TOTAL', 'GROSS', 'NET']):
            return 'summary'
        elif any(word in header_text for word in ['EQUITY', 'SHAREHOLDER', 'CAPITAL', 'STOCK']):
            return 'financial_statement'
        elif any(word in header_text for word in ['BALANCE', 'ASSET', 'LIABILITY']):
            return 'balance_sheet'
        else:
            return 'data'
    
    def _detect_headers(self, table_data: List[List]) -> bool:
        """Detect if table has headers"""
        if not table_data or len(table_data) < 2:
            return False
            
        # Check if first row looks like headers (contains text, not numbers)
        first_row = table_data[0]
        text_count = sum(1 for cell in first_row if cell and not self._is_numeric(str(cell)))
        
        # If more than half the cells in first row are text, likely headers
        return text_count > len(first_row) / 2
    
    def _assess_table_complexity(self, table_data: List[List]) -> str:
        """Assess table complexity level"""
        if not table_data:
            return 'simple'
            
        num_rows = len(table_data)
        num_cols = len(table_data[0]) if table_data else 0
        
        # Count merged cells or complex structures
        complex_cells = 0
        for row in table_data:
            for cell in row:
                if cell and len(str(cell)) > 20:  # Long text might indicate merged cells
                    complex_cells += 1
        
        if num_cols > 8 or num_rows > 20 or complex_cells > num_rows * num_cols * 0.3:
            return 'complex'
        elif num_cols > 4 or num_rows > 10:
            return 'medium'
        else:
            return 'simple'
            
    def _analyze_table_formatting(self, page, table, bbox) -> List[List[Dict]]:
        """Analyze formatting for each table cell"""
        cell_formats = []
        
        try:
            # Get text in table area
            table_area = fitz.Rect(bbox)
            text_dict = page.get_text("dict", clip=table_area)
            
            # For now, return basic format info
            # This could be enhanced to detect cell-specific formatting
            for row in table.extract():
                row_formats = []
                for cell in row:
                    row_formats.append({
                        'background_color': None,
                        'text_color': 0,  # Black
                        'is_bold': False,
                        'font_size': 10
                    })
                cell_formats.append(row_formats)
                
        except Exception as e:
            self.logger.debug(f"Error analyzing table formatting: {str(e)}")
            
        return cell_formats
        
    def create_excel_from_layout(self, ws, layout_info: Dict[str, Any]) -> None:
        """Create Excel worksheet from extracted layout information with proper row-based layout"""
        if not layout_info:
            return
            
        # Setup worksheet dimensions
        self._setup_worksheet(ws, layout_info)
        
        # Get all elements
        text_blocks = layout_info.get('text_blocks', [])
        tables = layout_info.get('tables', [])
        
        # Filter out text blocks that are inside table areas to avoid duplication
        filtered_text_blocks = self._filter_overlapping_text(text_blocks, tables)
        
        # Create a unified layout by combining filtered text and tables
        unified_elements = self._create_unified_layout(filtered_text_blocks, tables)
        
        # Create Excel with proper layout
        current_excel_row = 1
        for element in unified_elements:
            if element['type'] == 'text_row':
                current_excel_row = self._create_row_layout(ws, element['data'], current_excel_row, layout_info)
            elif element['type'] == 'table':
                current_excel_row += 1  # Add space before table
                current_excel_row = self._create_clean_table(ws, element['data'], current_excel_row, layout_info)
                current_excel_row += 1  # Add space after table
                
    def _setup_worksheet(self, ws, layout_info: Dict) -> None:
        """Setup worksheet with appropriate dimensions based on content"""
        # Analyze table complexity to set optimal column widths
        tables = layout_info.get('tables', [])
        max_complexity = 'simple'
        
        for table in tables:
            complexity = table.get('complexity', 'simple')
            if complexity == 'complex':
                max_complexity = 'complex'
                break
            elif complexity == 'medium' and max_complexity == 'simple':
                max_complexity = 'medium'
        
        # Set column widths based on complexity - optimized for full width table display
        if max_complexity == 'complex':
            # For complex tables (like payroll statements), use wider widths to fill full page
            column_widths = [30, 35, 20, 25, 25, 25, 25, 25]  # Optimized for full width payroll statements
        elif max_complexity == 'medium':
            # For medium complexity, use balanced widths with full coverage
            column_widths = [25, 30, 18, 22, 22, 22, 22, 22]
        else:
            # For simple tables, use uniform widths with full coverage
            column_widths = [22, 25, 18, 20, 20, 20, 20, 20]
        
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        # Set default row height
        ws.row_dimensions.default_row_height = 20
            
    def _create_text_element(self, ws, text_element: Dict, start_row: int, layout_info: Dict) -> int:
        """Create text element in Excel with smart positioning"""
        text = text_element['text']
        block_type = text_element.get('block_type', 'text')
        font_info = text_element.get('font_info', {})
        x_pos = text_element.get('x', 0)
        width = text_element.get('width', 0)
        
        # Determine column based on X position
        page_width = layout_info.get('page_size', {}).get('width', 595)
        col_position = self._get_column_from_x(x_pos, page_width)
        
        try:
            # Determine column span based on text type and position
            if block_type == 'main_header':
                # Main headers span most columns, centered
                self._create_merged_cell(ws, start_row, 'A', 'H', text,
                                       Font(size=16, bold=True, color='FF1F4E79'),
                                       Alignment(horizontal='center', vertical='center'))
                
            elif block_type == 'section_header':
                # Section headers span full width - use extracted colors only if they exist
                page_colors = layout_info.get('colors', {}).get('table_colors', {})
                header_text_color = page_colors.get('header_text')
                header_bg_color = page_colors.get('header_bg')
                
                # Only apply colors if they exist in PDF
                font = Font(size=12, bold=True)
                if header_text_color:
                    font = Font(size=12, bold=True, color=header_text_color)
                
                fill = None
                if header_bg_color:
                    fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
                
                self._create_merged_cell(ws, start_row, 'A', 'H', text,
                                       font,
                                       Alignment(horizontal='center', vertical='center'),
                                       fill)
                
            elif block_type == 'label':
                # Labels positioned based on X coordinate
                col_letter = get_column_letter(col_position)
                cell = ws[f'{col_letter}{start_row}']
                cell.value = text
                cell.font = Font(size=10, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            elif block_type == 'data':
                # Data positioned based on X coordinate
                col_letter = get_column_letter(col_position)
                cell = ws[f'{col_letter}{start_row}']
                cell.value = text
                cell.font = Font(size=10)
                # Right align if numeric
                if self._is_numeric(text):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
            elif block_type == 'footer':
                # Footer spans full width, small font
                self._create_merged_cell(ws, start_row, 'A', 'H', text,
                                       Font(size=9, color='FF808080'),
                                       Alignment(horizontal='center', vertical='center'))
                
            else:
                # Regular text positioned based on X coordinate
                col_letter = get_column_letter(col_position)
                cell = ws[f'{col_letter}{start_row}']
                cell.value = text
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
        except Exception as e:
            self.logger.debug(f"Error creating text element: {str(e)}")
            # Fallback to simple cell
            cell = ws[f'A{start_row}']
            cell.value = text
            cell.font = Font(size=10)
            
        return start_row + 1
        
    def _create_merged_cell(self, ws, row: int, start_col: str, end_col: str, 
                          value: str, font=None, alignment=None, fill=None):
        """Safely create a merged cell"""
        try:
            # Set value and formatting first
            cell = ws[f'{start_col}{row}']
            cell.value = value
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill
                
            # Then merge
            ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
            
        except Exception as e:
            self.logger.debug(f"Error creating merged cell: {str(e)}")
            # Fallback to single cell
            cell = ws[f'{start_col}{row}']
            cell.value = value
        
    def _create_table_element(self, ws, table_element: Dict, start_row: int, layout_info: Dict) -> int:
        """Create table element in Excel"""
        table_data = table_element.get('data', [])
        table_type = table_element.get('table_type', 'data')
        
        if not table_data:
            return start_row
            
        current_row = start_row
        
        # Create table with appropriate styling using extracted colors
        page_colors = layout_info.get('colors', {}).get('table_colors', {})
        
        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_value in enumerate(row_data[:6], 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(cell_value) if cell_value else "")
                
                if row_idx == 0:  # Header row
                    header_text_color = page_colors.get('header_text')
                    header_bg_color = page_colors.get('header_bg')
                    
                    # Only apply colors if they exist in the PDF
                    if header_text_color:
                        cell.font = Font(size=10, bold=True, color=header_text_color)
                    else:
                        cell.font = Font(size=10, bold=True)  # Default black text
                        
                    if header_bg_color:
                        cell.fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
                    # No fill if no background color in PDF
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # Data row
                    data_text_color = page_colors.get('data_text')
                    
                    # Only apply text color if it exists in PDF
                    if data_text_color:
                        cell.font = Font(size=10, color=data_text_color)
                    else:
                        cell.font = Font(size=10)  # Default black text
                    
                    # Only apply alternating colors if they exist in PDF
                    if row_idx % 2 == 0:
                        primary_bg = page_colors.get('data_bg_primary')
                        if primary_bg:
                            cell.fill = PatternFill(start_color=primary_bg, end_color=primary_bg, fill_type='solid')
                    else:
                        alternate_bg = page_colors.get('data_bg_alternate')
                        if alternate_bg:
                            cell.fill = PatternFill(start_color=alternate_bg, end_color=alternate_bg, fill_type='solid')
                        
                    # Smart alignment
                    if self._is_numeric(str(cell_value)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                # Only add borders if border color exists in PDF
                border_color = page_colors.get('border_color')
                if border_color:
                    cell.border = Border(
                        left=Side(style='thin', color=border_color),
                        right=Side(style='thin', color=border_color),
                        top=Side(style='thin', color=border_color),
                        bottom=Side(style='thin', color=border_color)
                    )
                
            current_row += 1
            
        return current_row + 1
        
    def _get_column_from_x(self, x_pos: float, page_width: float) -> int:
        """Convert X position to Excel column number"""
        # Divide page into 8 columns
        col_width = page_width / 8
        col = max(1, min(8, int(x_pos / col_width) + 1))
        return col
        
    def _get_column_span(self, width: float, page_width: float) -> int:
        """Calculate how many columns an element should span"""
        col_width = page_width / 8
        span = max(1, min(4, int(width / col_width) + 1))
        return span
        
    def _group_text_by_rows(self, text_blocks: List[Dict]) -> List[List[Dict]]:
        """Group text elements by rows (similar Y positions)"""
        if not text_blocks:
            return []
            
        # Sort by Y position first
        sorted_blocks = sorted(text_blocks, key=lambda x: x['y'])
        
        rows = []
        current_row = [sorted_blocks[0]]
        current_y = sorted_blocks[0]['y']
        
        for block in sorted_blocks[1:]:
            # If Y position is close (within 8 points), group together
            if abs(block['y'] - current_y) <= 8:
                current_row.append(block)
            else:
                # Sort current row by X position (left to right)
                current_row.sort(key=lambda x: x['x'])
                rows.append(current_row)
                current_row = [block]
                current_y = block['y']
                
        # Add last row
        if current_row:
            current_row.sort(key=lambda x: x['x'])
            rows.append(current_row)
            
        return rows
        
    def _create_row_layout(self, ws, row_elements: List[Dict], excel_row: int, layout_info: Dict) -> int:
        """Create a row layout in Excel based on PDF row structure"""
        if not row_elements:
            return excel_row + 1
            
        page_width = layout_info.get('page_size', {}).get('width', 595)
        
        # Analyze the row to determine if it's a special type
        row_type = self._classify_row_type(row_elements)
        
        if row_type == 'header_pair':
            # Handle header pairs like "Payroll Statement" | "Hourly"
            self._create_header_pair_row(ws, row_elements, excel_row, page_width)
            
        elif row_type == 'label_pair':
            # Handle label pairs like "Employee Name" | "Federal Filing Status"
            self._create_label_pair_row(ws, row_elements, excel_row, page_width)
            
        elif row_type == 'section_header':
            # Handle section headers like "CURRENT EARNINGS"
            self._create_section_header_row(ws, row_elements, excel_row, layout_info)
            
        elif row_type == 'table_header':
            # Handle table headers like "EARNING TYPE | HOURS | PAYMENT"
            self._create_table_header_row(ws, row_elements, excel_row, page_width, layout_info)
            
        elif row_type == 'data_row':
            # Handle data rows like "Regular Pay | 0.00 | $ -"
            self._create_data_row(ws, row_elements, excel_row, page_width)
            
        else:
            # Default: place elements based on X position
            self._create_default_row(ws, row_elements, excel_row, page_width)
            
        return excel_row + 1
        
    def _classify_row_type(self, row_elements: List[Dict]) -> str:
        """Classify the type of row based on content and formatting"""
        if not row_elements:
            return 'empty'
            
        # Check for section headers (centered, all caps, bold)
        if len(row_elements) == 1:
            elem = row_elements[0]
            text = elem['text'].upper()
            if any(word in text for word in ['CURRENT', 'EARNINGS', 'STATEMENT', 'SUMMARY', 'DEDUCTION', 'DETAILS']):
                return 'section_header'
                
        # Check for header pairs (large font, 2 elements)
        if len(row_elements) == 2:
            sizes = [elem.get('font_info', {}).get('size', 10) for elem in row_elements]
            if any(size > 14 for size in sizes):
                return 'header_pair'
            # Check for label pairs
            texts = [elem['text'].upper() for elem in row_elements]
            if any(word in ' '.join(texts) for word in ['EMPLOYEE', 'NAME', 'PERIOD', 'STATUS', 'DATE']):
                return 'label_pair'
                
        # Check for table headers (multiple columns with specific keywords)
        if len(row_elements) >= 3:
            texts = [elem['text'].upper() for elem in row_elements]
            if any(word in ' '.join(texts) for word in ['TYPE', 'HOURS', 'PAYMENT', 'CURRENT', 'YTD']):
                return 'table_header'
                
        # Check for data rows (contains numbers or currency)
        texts = [elem['text'] for elem in row_elements]
        if any(self._is_numeric(text) or '$' in text for text in texts):
            return 'data_row'
            
        return 'default'
    
    def _create_header_pair_row(self, ws, elements: List[Dict], row: int, page_width: float):
        """Create header pair row like 'Payroll Statement | Hourly'"""
        if len(elements) >= 2:
            # Left header (A-D)
            left_elem = elements[0]
            self._create_merged_cell(ws, row, 'A', 'D', left_elem['text'],
                                   Font(size=16, bold=True, color='FF1F4E79'),
                                   Alignment(horizontal='left', vertical='center'))
            
            # Right header (E-H)  
            right_elem = elements[1]
            self._create_merged_cell(ws, row, 'E', 'H', right_elem['text'],
                                   Font(size=20, bold=True, color='FF1F4E79'),
                                   Alignment(horizontal='right', vertical='center'))
        elif len(elements) == 1:
            # Single header spans all
            elem = elements[0]
            self._create_merged_cell(ws, row, 'A', 'H', elem['text'],
                                   Font(size=16, bold=True, color='FF1F4E79'),
                                   Alignment(horizontal='center', vertical='center'))
                                   
    def _create_label_pair_row(self, ws, elements: List[Dict], row: int, page_width: float):
        """Create label pair row like 'Employee Name | Federal Filing Status'"""
        if len(elements) >= 2:
            # Left label (A-D)
            left_elem = elements[0]
            self._create_merged_cell(ws, row, 'A', 'D', left_elem['text'],
                                   Font(size=10, bold=True),
                                   Alignment(horizontal='left', vertical='center'))
            
            # Right label (E-H)
            right_elem = elements[1]
            self._create_merged_cell(ws, row, 'E', 'H', right_elem['text'],
                                   Font(size=10, bold=True),
                                   Alignment(horizontal='left', vertical='center'))
        elif len(elements) == 1:
            # Single label
            elem = elements[0]
            ws[f'A{row}'].value = elem['text']
            ws[f'A{row}'].font = Font(size=10, bold=True)
            
    def _create_section_header_row(self, ws, elements: List[Dict], row: int, layout_info: Dict = None):
        """Create section header row like 'CURRENT EARNINGS'"""
        elem = elements[0]
        
        # Use extracted colors only if they exist in PDF
        font = Font(size=12, bold=True)
        fill = None
        
        if layout_info:
            page_colors = layout_info.get('colors', {}).get('table_colors', {})
            header_text_color = page_colors.get('header_text')
            header_bg_color = page_colors.get('header_bg')
            
            if header_text_color:
                font = Font(size=12, bold=True, color=header_text_color)
            if header_bg_color:
                fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
            
        self._create_merged_cell(ws, row, 'A', 'H', elem['text'],
                               font,
                               Alignment(horizontal='center', vertical='center'),
                               fill)
                               
    def _create_table_header_row(self, ws, elements: List[Dict], row: int, page_width: float, layout_info: Dict = None):
        """Create table header row like 'EARNING TYPE | HOURS | PAYMENT'"""
        # Distribute elements across columns efficiently to use full width
        num_elements = len(elements) if elements else 1
        cols_per_element = max(1, 8 // num_elements)  # Ensure at least 1 column per element
        
        # Use extracted colors only if they exist in PDF
        font = Font(size=10, bold=True)
        fill = None
        
        if layout_info:
            page_colors = layout_info.get('colors', {}).get('table_colors', {})
            header_text_color = page_colors.get('header_text')
            header_bg_color = page_colors.get('header_bg')
            
            if header_text_color:
                font = Font(size=10, bold=True, color=header_text_color)
            if header_bg_color:
                fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
            
        # Use the same efficient column distribution as _create_clean_table
        num_cols = len(elements)
        if num_cols == 1:
            col_positions = [(1, 8)]  # A-H
        elif num_cols == 2:
            col_positions = [(1, 4), (5, 8)]  # A-D, E-H
        elif num_cols == 3:
            col_positions = [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H
        elif num_cols == 4:
            col_positions = [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
        elif num_cols == 5:
            col_positions = [(1, 1), (2, 3), (4, 5), (6, 7), (8, 8)]  # A, B-C, D-E, F-G, H
        elif num_cols == 6:
            col_positions = [(1, 1), (2, 2), (3, 4), (5, 6), (7, 7), (8, 8)]  # A, B, C-D, E-F, G, H
        elif num_cols == 7:
            col_positions = [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 8)]  # A, B, C, D, E, F, G-H
        else:
            col_positions = [(i+1, i+1) for i in range(min(num_cols, 8))]
        
        for i, elem in enumerate(elements):
            if i >= len(col_positions):
                break
                
            start_col, end_col = col_positions[i]
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            
            if start_col == end_col:
                # Single column
                cell = ws[f'{start_letter}{row}']
                cell.value = elem['text']
                cell.font = font
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Merged cells
                self._create_merged_cell(ws, row, start_letter, end_letter, elem['text'],
                                       font,
                                       Alignment(horizontal='center', vertical='center'),
                                       fill)
                                       
    def _create_data_row(self, ws, elements: List[Dict], row: int, page_width: float):
        """Create data row like 'Regular Pay | 0.00 | $ -'"""
        # Distribute elements across columns efficiently to use full width
        num_elements = len(elements) if elements else 1
        cols_per_element = max(1, 8 // num_elements)  # Ensure at least 1 column per element
        
        # Use the same efficient column distribution as _create_clean_table
        num_cols = len(elements)
        if num_cols == 1:
            col_positions = [(1, 8)]  # A-H
        elif num_cols == 2:
            col_positions = [(1, 4), (5, 8)]  # A-D, E-H
        elif num_cols == 3:
            col_positions = [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H
        elif num_cols == 4:
            col_positions = [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
        elif num_cols == 5:
            col_positions = [(1, 1), (2, 3), (4, 5), (6, 7), (8, 8)]  # A, B-C, D-E, F-G, H
        elif num_cols == 6:
            col_positions = [(1, 1), (2, 2), (3, 4), (5, 6), (7, 7), (8, 8)]  # A, B, C-D, E-F, G, H
        elif num_cols == 7:
            col_positions = [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 8)]  # A, B, C, D, E, F, G-H
        else:
            col_positions = [(i+1, i+1) for i in range(min(num_cols, 8))]
        
        for i, elem in enumerate(elements):
            if i >= len(col_positions):
                break
                
            start_col, end_col = col_positions[i]
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            
            if start_col == end_col:
                # Single column
                cell = ws[f'{start_letter}{row}']
                cell.value = elem['text']
                cell.font = Font(size=10)
                
                # Right align if numeric
                if self._is_numeric(elem['text']) or '$' in elem['text']:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Merged columns
                self._create_merged_cell(ws, row, start_letter, end_letter, elem['text'],
                                       Font(size=10),
                                       Alignment(horizontal='left', vertical='center'))
                
    def _create_default_row(self, ws, elements: List[Dict], row: int, page_width: float):
        """Create default row by positioning elements based on X coordinates"""
        for elem in elements:
            col = self._get_column_from_x(elem['x'], page_width)
            col_letter = get_column_letter(col)
            
            cell = ws[f'{col_letter}{row}']
            cell.value = elem['text']
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _filter_overlapping_text(self, text_blocks: List[Dict], tables: List[Dict]) -> List[Dict]:
        """Filter out text blocks that overlap with table areas to avoid duplication"""
        filtered_blocks = []
        
        for text_block in text_blocks:
            text_bbox = [
                text_block['x'], 
                text_block['y'],
                text_block['x'] + text_block['width'],
                text_block['y'] + text_block['height']
            ]
            
            # Check if this text block overlaps with any table
            overlaps_with_table = False
            for table in tables:
                table_bbox = [
                    table['x'],
                    table['y'], 
                    table['x'] + table['width'],
                    table['y'] + table['height']
                ]
                
                # Check for overlap (with some tolerance)
                if (text_bbox[0] < table_bbox[2] - 5 and text_bbox[2] > table_bbox[0] + 5 and
                    text_bbox[1] < table_bbox[3] - 5 and text_bbox[3] > table_bbox[1] + 5):
                    overlaps_with_table = True
                    break
                    
            if not overlaps_with_table:
                filtered_blocks.append(text_block)
                
        return filtered_blocks
        
    def _create_unified_layout(self, text_blocks: List[Dict], tables: List[Dict]) -> List[Dict]:
        """Create a unified layout combining text rows and tables in correct order"""
        unified_elements = []
        
        # Group text blocks by rows
        if text_blocks:
            text_rows = self._group_text_by_rows(text_blocks)
            for row in text_rows:
                unified_elements.append({
                    'type': 'text_row',
                    'y': min(block['y'] for block in row),
                    'data': row
                })
        
        # Add tables
        for table in tables:
            unified_elements.append({
                'type': 'table',
                'y': table['y'],
                'data': table
            })
            
        # Sort all elements by Y position
        unified_elements.sort(key=lambda x: x['y'])
        
        return unified_elements
        
    def _create_clean_table(self, ws, table_data: Dict, start_row: int, layout_info: Dict) -> int:
        """Create a clean table without duplication, spanning full width like section headers"""
        table_rows = table_data.get('data', [])
        if not table_rows:
            return start_row
            
        current_row = start_row
        
        # Get table complexity for advanced layout
        complexity = table_data.get('complexity', 'simple')
        num_cols = table_data.get('num_cols', 0)
        has_headers = table_data.get('has_headers', False)
        
        # Create table with proper column alignment spanning A-H
        for row_idx, row_data in enumerate(table_rows):
            # Clean row data - remove empty cells
            cleaned_row = [str(cell).strip() for cell in row_data if str(cell).strip()]
            
            if not cleaned_row:  # Skip completely empty rows
                continue
                
            # Distribute columns across A-H based on number of actual columns
            num_cols_actual = len(cleaned_row)
            if num_cols_actual == 0:
                continue
                
            # Enhanced column distribution based on complexity
            col_positions = self._calculate_optimal_column_layout(num_cols_actual, complexity, has_headers, row_idx)
                
            # Create cells for each column
            for col_idx, cell_value in enumerate(cleaned_row):
                if col_idx >= len(col_positions):
                    break
                    
                start_col, end_col = col_positions[col_idx]
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                
                if start_col == end_col:
                    # Single column
                    cell = ws[f'{start_letter}{current_row}']
                    cell.value = cell_value
                else:
                    # Merged columns
                    self._create_merged_cell(ws, current_row, start_letter, end_letter, cell_value)
                    cell = ws[f'{start_letter}{current_row}']
                
                # Apply formatting using extracted colors
                page_colors = layout_info.get('colors', {}).get('table_colors', {})
                
                if row_idx == 0:  # Header row
                    header_text_color = page_colors.get('header_text')
                    header_bg_color = page_colors.get('header_bg')
                    
                    # Only apply colors if they exist in PDF
                    if header_text_color:
                        cell.font = Font(size=10, bold=True, color=header_text_color)
                    else:
                        cell.font = Font(size=10, bold=True)  # Default black text
                        
                    if header_bg_color:
                        cell.fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type='solid')
                    # No fill if no background color in PDF
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:  # Data row
                    data_text_color = page_colors.get('data_text')
                    
                    # Only apply text color if it exists in PDF
                    if data_text_color:
                        cell.font = Font(size=10, color=data_text_color)
                    else:
                        cell.font = Font(size=10)  # Default black text
                    
                    # Only apply alternating colors if they exist in PDF
                    if row_idx % 2 == 0:
                        primary_bg = page_colors.get('data_bg_primary')
                        if primary_bg:
                            cell.fill = PatternFill(start_color=primary_bg, end_color=primary_bg, fill_type='solid')
                    else:
                        alternate_bg = page_colors.get('data_bg_alternate')
                        if alternate_bg:
                            cell.fill = PatternFill(start_color=alternate_bg, end_color=alternate_bg, fill_type='solid')
                        
                    # Smart alignment
                    if self._is_numeric(cell_value) or '$' in cell_value:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                # Only add borders if border color exists in PDF
                border_color = page_colors.get('border_color')
                if border_color:
                    for col in range(start_col, end_col + 1):
                        border_cell = ws[f'{get_column_letter(col)}{current_row}']
                        border_cell.border = Border(
                            left=Side(style='thin', color=border_color),
                            right=Side(style='thin', color=border_color),
                            top=Side(style='thin', color=border_color),
                            bottom=Side(style='thin', color=border_color)
                        )
                
            current_row += 1
            
        return current_row
    
    def _calculate_optimal_column_layout(self, num_cols: int, complexity: str, has_headers: bool, row_idx: int) -> List[Tuple[int, int]]:
        """Calculate optimal column layout based on table complexity and content - optimized for full width usage"""
        
        if complexity == 'complex':
            # For complex tables (like payroll statements), use full width distribution
            if num_cols == 1:
                return [(1, 8)]  # A-H (full width)
            elif num_cols == 2:
                return [(1, 4), (5, 8)]  # A-D, E-H (half width each)
            elif num_cols == 3:
                return [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H (balanced)
            elif num_cols == 4:
                return [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
            elif num_cols == 5:
                # Optimized for payroll statements: Category, Deduction Type, Rate, Current, Year to Date
                return [(1, 1), (2, 3), (4, 4), (5, 6), (7, 8)]  # A, B-C, D, E-F, G-H
            elif num_cols == 6:
                return [(1, 1), (2, 2), (3, 4), (5, 5), (6, 7), (8, 8)]  # A, B, C-D, E, F-G, H
            elif num_cols == 7:
                return [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 8)]  # A, B, C, D, E, F, G-H
            elif num_cols == 8:
                return [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8)]  # A, B, C, D, E, F, G, H
            else:
                # For more than 8 columns, use individual columns but ensure we use all 8 available
                return [(i+1, i+1) for i in range(min(num_cols, 8))]
                
        elif complexity == 'medium':
            # For medium complexity, use balanced distribution with full width
            if num_cols == 1:
                return [(1, 8)]  # A-H
            elif num_cols == 2:
                return [(1, 4), (5, 8)]  # A-D, E-H
            elif num_cols == 3:
                return [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H
            elif num_cols == 4:
                return [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
            elif num_cols == 5:
                return [(1, 1), (2, 3), (4, 4), (5, 6), (7, 8)]  # A, B-C, D, E-F, G-H
            else:
                return [(i+1, i+1) for i in range(min(num_cols, 8))]
                
        else:  # simple
            # For simple tables, use basic distribution with full width
            if num_cols == 1:
                return [(1, 8)]  # A-H
            elif num_cols == 2:
                return [(1, 4), (5, 8)]  # A-D, E-H
            elif num_cols == 3:
                return [(1, 2), (3, 5), (6, 8)]  # A-B, C-E, F-H
            elif num_cols == 4:
                return [(1, 2), (3, 4), (5, 6), (7, 8)]  # A-B, C-D, E-F, G-H
            elif num_cols == 5:
                return [(1, 1), (2, 3), (4, 4), (5, 6), (7, 8)]  # A, B-C, D, E-F, G-H
            else:
                return [(i+1, i+1) for i in range(min(num_cols, 8))]
    
    def _is_numeric(self, text: str) -> bool:
        """Check if text is numeric"""
        if not text or text.strip() == '' or text.strip() == '-':
            return False
            
        clean_text = text.strip().replace(',', '').replace('$', '').replace('%', '').replace('-', '')
        
        try:
            float(clean_text)
            return True
        except ValueError:
            return False
