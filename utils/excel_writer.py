#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Writer Module
==================

Writes extracted PDF data to Excel files with proper formatting and multiple sheets.
Supports various output formats and styling options.
"""

import os
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

PANDAS_AVAILABLE = False
EXCEL_LIBS_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    print("pandas not available - using basic Excel writing")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_LIBS_AVAILABLE = True
    
    # Only import dataframe_to_rows if pandas is available
    if PANDAS_AVAILABLE:
        from openpyxl.utils.dataframe import dataframe_to_rows
        
except ImportError as e:
    EXCEL_LIBS_AVAILABLE = False
    print(f"Excel libraries not available: {e}")
    print("Install with: pip install openpyxl")


class ColorManager:
    """
    Manages colors extracted from PDF and provides Excel-compatible colors
    """
    def __init__(self, pdf_colors: Dict[str, Any] = None):
        """Initialize with PDF colors"""
        self.pdf_colors = pdf_colors or {}
        # Professional color scheme as fallback
        self._default_colors = {
            'header_bg': 'FF1E3A8A',      # Professional blue
            'header_text': 'FFFFFFFF',    # White
            'title_text': 'FF1F2937',     # Dark gray
            'section_text': 'FF374151',   # Medium gray
            'table_title_bg': 'FFF1F5F9', # Very light blue-gray
            'table_title_text': 'FF475569', # Medium gray
            'data_text': 'FF1F2937',      # Dark gray
            'accent_color': 'FF3B82F6'    # Bright blue
        }
    
    def get_header_colors(self) -> tuple:
        """Get header background and text colors"""
        table_colors = self.pdf_colors.get('table_colors', {})
        bg_color = table_colors.get('header_bg', self._default_colors['header_bg'])
        text_color = table_colors.get('header_text', self._default_colors['header_text'])
        
        # Ensure good contrast
        if self._is_light_color(bg_color):
            text_color = self._default_colors['title_text']  # Dark text on light bg
        
        return bg_color, text_color
    
    def get_title_color(self) -> str:
        """Get title text color"""
        table_colors = self.pdf_colors.get('table_colors', {})
        data_text = table_colors.get('data_text')
        
        if data_text and self._is_readable_color(data_text):
            return data_text
        
        # Fallback to professional color
        return self._default_colors['title_text']
    
    def get_section_color(self) -> str:
        """Get section header color"""
        table_colors = self.pdf_colors.get('table_colors', {})
        data_text = table_colors.get('data_text')
        
        if data_text and self._is_readable_color(data_text):
            return data_text
            
        return self._default_colors['section_text']
    
    def get_table_title_colors(self) -> tuple:
        """Get table title background and text colors"""
        table_colors = self.pdf_colors.get('table_colors', {})
        
        # Use subtle background color
        bg_color = table_colors.get('data_bg_alternate', self._default_colors['table_title_bg'])
        
        # Ensure background is light enough for table titles
        if not self._is_light_color(bg_color):
            bg_color = self._default_colors['table_title_bg']
        
        # Use readable text color
        text_color = table_colors.get('data_text', self._default_colors['table_title_text'])
        if not self._is_readable_color(text_color):
            text_color = self._default_colors['table_title_text']
        
        return bg_color, text_color
    
    def _is_light_color(self, hex_color: str) -> bool:
        """Check if color is light (suitable for backgrounds)"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)
                b = int(hex_color[6:8], 16)
                brightness = (r * 299 + g * 587 + b * 114) / 1000
                return brightness > 180
        except:
            pass
        return False
    
    def _is_readable_color(self, hex_color: str) -> bool:
        """Check if color is readable for text"""
        try:
            if len(hex_color) >= 8:
                r = int(hex_color[2:4], 16)
                g = int(hex_color[4:6], 16)
                b = int(hex_color[6:8], 16)
                brightness = (r * 299 + g * 587 + b * 114) / 1000
                return brightness < 150  # Dark enough to read
        except:
            pass
        return False


class ExcelWriter:
    """
    Excel file writer with formatting and multi-sheet support
    
    Features:
    - Multiple sheets for different data types
    - Professional formatting and styling
    - Auto-sizing columns
    - Table formatting with filters
    - Summary statistics
    """
    
    def __init__(self):
        """Initialize Excel writer"""
        self.logger = logging.getLogger(__name__)
        self.color_manager = None
        
        if not EXCEL_LIBS_AVAILABLE:
            self.logger.error("Excel libraries not available")
            
    def write_to_excel(self, pdf_data: Dict[str, Any], output_path: str) -> bool:
        """
        Write PDF data to Excel file
        
        Args:
            pdf_data (dict): Extracted PDF data
            output_path (str): Output Excel file path
            
        Returns:
            bool: True if successful
        """
        if not EXCEL_LIBS_AVAILABLE:
            return self._write_fallback_format(pdf_data, output_path)
            
        try:
            # Initialize color manager with PDF colors
            pdf_colors = {}
            if 'pages' in pdf_data and pdf_data['pages']:
                # Get colors from first page (assuming consistent colors)
                first_page = pdf_data['pages'][0]
                pdf_colors = first_page.get('colors', {})
            
            self.color_manager = ColorManager(pdf_colors)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                
            # Create sheets for different data types (summary first)
            self._create_summary_sheet(wb, pdf_data)
            self._create_text_sheet(wb, pdf_data)
            self._create_tables_sheet(wb, pdf_data)
            self._create_metadata_sheet(wb, pdf_data)
            self._create_images_sheet(wb, pdf_data)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Excel file created successfully: {output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error writing Excel file: {str(e)}")
            return False
            
    def _create_text_sheet(self, wb, pdf_data: Dict[str, Any]):
        """Create sheet for extracted text"""
        ws = wb.create_sheet("Document_Text")
        
        # Headers
        headers = ["Page", "Text Content", "Character Count", "Word Count", "Primary Font", "Font Size"]
        ws.append(headers)
        
        # Style headers
        self._style_header_row(ws, 1, len(headers))
        
        # Add text data
        text_data = pdf_data.get('text_data', [])
        for text_info in text_data:
            # Get primary font info
            fonts = text_info.get('fonts', [])
            primary_font = fonts[0] if fonts else {}
            
            row = [
                text_info.get('page', ''),
                text_info.get('text', ''),
                text_info.get('char_count', 0),
                text_info.get('word_count', 0),
                primary_font.get('name', ''),
                primary_font.get('size', '')
            ]
            ws.append(row)
            
        # Auto-size columns
        self._auto_size_columns(ws)
        
        # Format as table
        if len(text_data) > 0:
            self._format_as_table(ws, len(text_data) + 1, len(headers), "TextData")
            
    def _create_tables_sheet(self, wb: Workbook, pdf_data: Dict[str, Any]):
        """Create sheet for extracted tables"""
        ws = wb.create_sheet("Extracted_Tables")
        
        tables_data = pdf_data.get('tables', [])
        
        if not tables_data:
            ws.append(["No tables were found in the PDF"])
            return
            
        current_row = 1
        
        for i, table_info in enumerate(tables_data):
            # Table header
            table_title = f"Table {table_info.get('table_id', i+1)} (Page {table_info.get('page', 'Unknown')})"
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = table_title
            self._style_table_title(ws, current_row)
            current_row += 1
            
            # Table metadata
            metadata = [
                f"Extraction Method: {table_info.get('extraction_method', 'Unknown')}",
                f"Rows: {table_info.get('rows', 0)}",
                f"Columns: {table_info.get('columns', 0)}",
                f"Confidence: {table_info.get('confidence', 0):.2f}"
            ]
            
            for meta in metadata:
                ws[f'A{current_row}'] = meta
                current_row += 1
                
            current_row += 1  # Empty row
            
            # Table data
            table_data = table_info.get('data')
            if table_data:
                # Handle pandas DataFrame
                if PANDAS_AVAILABLE and hasattr(table_data, 'empty') and not table_data.empty:
                    # Add table headers
                    headers = [f"Col_{i+1}" for i in range(len(table_data.columns))]
                    ws.append(headers)
                    header_row = current_row
                    current_row += 1
                    
                    # Add table data
                    for _, row in table_data.iterrows():
                        ws.append([str(cell) for cell in row])
                        current_row += 1
                        
                    # Format table
                    if len(table_data) > 0:
                        self._format_as_table(ws, current_row - header_row, len(headers), f"Table_{i+1}", header_row)
                        
                # Handle basic list format
                elif isinstance(table_data, list) and table_data:
                    # Add table headers
                    headers = [f"Col_{i+1}" for i in range(len(table_data[0]) if table_data[0] else 0)]
                    if headers:
                        ws.append(headers)
                        header_row = current_row
                        current_row += 1
                        
                        # Add table data
                        for row in table_data:
                            ws.append([str(cell) if cell else "" for cell in row])
                            current_row += 1
                            
                        # Format table
                        if len(table_data) > 0:
                            self._format_as_table(ws, current_row - header_row, len(headers), f"Table_{i+1}", header_row)
                    
            current_row += 2  # Space between tables
            
        self._auto_size_columns(ws)
        
    def _create_metadata_sheet(self, wb: Workbook, pdf_data: Dict[str, Any]):
        """Create sheet for PDF metadata"""
        ws = wb.create_sheet("Document_Metadata")
        
        # Headers
        ws.append(["Property", "Value"])
        self._style_header_row(ws, 1, 2)
        
        # Add metadata
        metadata = pdf_data.get('metadata', {})
        
        # Standard metadata fields
        standard_fields = [
            ('Filename', pdf_data.get('filename', '')),
            ('Title', metadata.get('title', '')),
            ('Author', metadata.get('author', '')),
            ('Subject', metadata.get('subject', '')),
            ('Creator', metadata.get('creator', '')),
            ('Producer', metadata.get('producer', '')),
            ('Creation Date', metadata.get('creation_date', '')),
            ('Modification Date', metadata.get('modification_date', '')),
            ('Page Count', metadata.get('page_count', '')),
            ('Encrypted', metadata.get('encrypted', '')),
            ('File Size', self._format_file_size(metadata.get('file_size', 0)))
        ]
        
        for prop, value in standard_fields:
            ws.append([prop, str(value)])
            
        # Processing information
        ws.append(['', ''])  # Empty row
        ws.append(['Processing Information', ''])
        self._style_header_row(ws, ws.max_row, 2)
        
        processing_info = [
            ('Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Text Pages Processed', len(pdf_data.get('text_data', []))),
            ('Tables Found', len(pdf_data.get('tables', []))),
            ('Images Extracted', len(pdf_data.get('images', [])))
        ]
        
        for prop, value in processing_info:
            ws.append([prop, str(value)])
            
        self._auto_size_columns(ws)
        
    def _create_images_sheet(self, wb: Workbook, pdf_data: Dict[str, Any]):
        """Create sheet for image information"""
        ws = wb.create_sheet("Images_Information")
        
        images_data = pdf_data.get('images', [])
        
        if not images_data:
            ws.append(["No images were found in the PDF"])
            return
            
        # Headers
        headers = ["Image ID", "Filename", "Page", "Width", "Height", "Format", "File Size", "Color Mode", "Quality"]
        ws.append(headers)
        self._style_header_row(ws, 1, len(headers))
        
        # Add image data
        for img_info in images_data:
            row = [
                img_info.get('image_id', ''),
                img_info.get('filename', ''),
                img_info.get('page', ''),
                img_info.get('width', ''),
                img_info.get('height', ''),
                img_info.get('format', ''),
                self._format_file_size(img_info.get('file_size', 0)),
                img_info.get('mode', ''),
                img_info.get('estimated_quality', '')
            ]
            ws.append(row)
            
        # Format as table
        if len(images_data) > 0:
            self._format_as_table(ws, len(images_data) + 1, len(headers), "ImagesData")
            
        self._auto_size_columns(ws)
        
    def _create_summary_sheet(self, wb: Workbook, pdf_data: Dict[str, Any]):
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary")
        # Move to first position (openpyxl doesn't have move method)
        # We'll handle this by creating summary sheet first
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"PDF Conversion Summary: {pdf_data.get('filename', 'Unknown')}"
        self._style_title(ws, 'A1')
        
        # Summary statistics
        current_row = 3
        
        sections = [
            ("Document Information", [
                ("Filename", pdf_data.get('filename', '')),
                ("Processing Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                ("Total Pages", len(pdf_data.get('text_data', [])))
            ]),
            ("Text Extraction", [
                ("Pages with Text", len(pdf_data.get('text_data', []))),
                ("Total Characters", sum(t.get('char_count', 0) for t in pdf_data.get('text_data', []))),
                ("Total Words", sum(t.get('word_count', 0) for t in pdf_data.get('text_data', [])))
            ]),
            ("Tables", [
                ("Tables Found", len(pdf_data.get('tables', []))),
                ("Pages with Tables", len(set(t.get('page', 0) for t in pdf_data.get('tables', [])))),
                ("Total Table Rows", sum(t.get('rows', 0) for t in pdf_data.get('tables', [])))
            ]),
            ("Images", [
                ("Images Extracted", len(pdf_data.get('images', []))),
                ("Pages with Images", len(set(i.get('page', 0) for i in pdf_data.get('images', [])))),
                ("Total Image Size", self._format_file_size(sum(i.get('file_size', 0) for i in pdf_data.get('images', []))))
            ])
        ]
        
        for section_title, items in sections:
            # Section header
            ws[f'A{current_row}'] = section_title
            self._style_section_header(ws, current_row)
            current_row += 1
            
            # Section items
            for label, value in items:
                ws[f'A{current_row}'] = label
                ws[f'B{current_row}'] = str(value)
                current_row += 1
                
            current_row += 1  # Empty row between sections
            
        self._auto_size_columns(ws)
        
    def _style_header_row(self, ws, row_num: int, col_count: int):
        """Apply header styling to a row using extracted colors"""
        if self.color_manager:
            bg_color, text_color = self.color_manager.get_header_colors()
        else:
            bg_color, text_color = "FF366092", "FFFFFFFF"
            
        header_font = Font(bold=True, color=text_color)
        header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    def _style_title(self, ws, cell_ref: str):
        """Style title cell using extracted colors"""
        cell = ws[cell_ref]
        
        if self.color_manager:
            title_color = self.color_manager.get_title_color()
        else:
            title_color = "FF2F4F4F"
            
        cell.font = Font(size=16, bold=True, color=title_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    def _style_section_header(self, ws, row_num: int):
        """Style section header using extracted colors"""
        cell = ws.cell(row=row_num, column=1)
        
        if self.color_manager:
            section_color = self.color_manager.get_section_color()
        else:
            section_color = "FF2F4F4F"
            
        cell.font = Font(bold=True, size=12, color=section_color)
        
    def _style_table_title(self, ws, row_num: int):
        """Style table title using extracted colors"""
        cell = ws.cell(row=row_num, column=1)
        
        if self.color_manager:
            bg_color, text_color = self.color_manager.get_table_title_colors()
        else:
            bg_color, text_color = "FFE6E6FA", "FF2F4F4F"
            
        cell.font = Font(bold=True, color=text_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
    def _format_as_table(self, ws, row_count: int, col_count: int, table_name: str, start_row: int = 1):
        """Format range as Excel table"""
        try:
            # Get column letter safely
            from openpyxl.utils import get_column_letter
            end_col_letter = get_column_letter(col_count)
            table_range = f"A{start_row}:{end_col_letter}{start_row + row_count - 1}"
            
            table = Table(displayName=table_name, ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            
        except Exception as e:
            self.logger.warning(f"Could not format as table: {str(e)}")
            
    def _auto_size_columns(self, ws):
        """Auto-size all columns in worksheet"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                        
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Could not auto-size columns: {str(e)}")
            
    def _format_file_size(self, size_bytes: int) -> str:
        """Format file size in human-readable format"""
        if size_bytes == 0:
            return "0 B"
            
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
            
        return f"{size_bytes:.1f} {size_names[i]}"
        
    def _write_fallback_format(self, pdf_data: Dict[str, Any], output_path: str) -> bool:
        """Write data in fallback format when Excel libraries not available"""
        try:
            # Create text-based output
            fallback_path = output_path.replace('.xlsx', '_fallback.txt')
            
            with open(fallback_path, 'w', encoding='utf-8') as f:
                f.write(f"PDF Conversion Results\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"Document: {pdf_data.get('filename', 'Unknown')}\n")
                f.write(f"Processing Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                # Text data
                f.write("EXTRACTED TEXT\n")
                f.write("-" * 30 + "\n")
                for text_info in pdf_data.get('text_data', []):
                    f.write(f"Page {text_info.get('page', 'Unknown')}:\n")
                    f.write(f"{text_info.get('text', '')}\n\n")
                    
                # Tables
                f.write("\nEXTRACTED TABLES\n")
                f.write("-" * 30 + "\n")
                for table_info in pdf_data.get('tables', []):
                    f.write(f"Table {table_info.get('table_id', 'Unknown')} from Page {table_info.get('page', 'Unknown')}:\n")
                    f.write(f"Method: {table_info.get('extraction_method', 'Unknown')}\n")
                    f.write(f"Size: {table_info.get('rows', 0)} rows x {table_info.get('columns', 0)} columns\n")
                    if 'csv_data' in table_info:
                        f.write(f"{table_info['csv_data']}\n")
                    f.write("\n")
                    
                # Images
                f.write("\nEXTRACTED IMAGES\n")
                f.write("-" * 30 + "\n")
                for img_info in pdf_data.get('images', []):
                    f.write(f"Image: {img_info.get('filename', 'Unknown')}\n")
                    f.write(f"Page: {img_info.get('page', 'Unknown')}\n")
                    f.write(f"Size: {img_info.get('width', 0)} x {img_info.get('height', 0)}\n")
                    f.write(f"Format: {img_info.get('format', 'Unknown')}\n\n")
                    
            self.logger.info(f"Fallback format file created: {fallback_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error writing fallback format: {str(e)}")
            return False
